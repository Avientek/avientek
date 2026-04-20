"""Ghost Voucher diagnostics & export.

A "ghost voucher" is a submitted document (docstatus=1) that has no
corresponding GL Entry (for financial docs) or Stock Ledger Entry
(for inventory docs). They look submitted in list view but are
invisible to accounting/stock reports.

This module scans for them across all finance/inventory doctypes and
can export the full list to Excel for sharing with the accounts team.
"""

import io

import frappe
from frappe.utils import flt, getdate


# Doctypes that post to GL Entry (financial)
GL_DOCTYPES = [
	("Sales Invoice", "grand_total"),
	("Purchase Invoice", "grand_total"),
	("Journal Entry", "total_debit"),
	("Payment Entry", "paid_amount"),
	("Expense Claim", "total_sanctioned_amount"),
]

# Doctypes that post to Stock Ledger Entry (inventory)
SLE_DOCTYPES = [
	("Delivery Note", "grand_total"),
	("Stock Entry", "total_amount"),
	("Purchase Receipt", "grand_total"),
]


def _collect_ghosts():
	"""Return list of dicts — one per ghost voucher."""
	rows = []

	# Financial docs — missing GL Entry
	for dt, amount_field in GL_DOCTYPES:
		submitted = frappe.db.sql(
			f"""
			SELECT name, company, posting_date, owner, `{amount_field}` AS amount
			FROM `tab{dt}`
			WHERE docstatus = 1
			""",
			as_dict=True,
		)
		for d in submitted:
			has_gl = frappe.db.exists(
				"GL Entry",
				{"voucher_type": dt, "voucher_no": d.name, "is_cancelled": 0},
			)
			if not has_gl:
				rows.append(
					{
						"doctype": dt,
						"voucher_no": d.name,
						"company": d.company,
						"posting_date": d.posting_date,
						"amount": flt(d.amount),
						"owner": d.owner,
						"missing": "GL Entry",
						"category": "Financial",
					}
				)

	# Inventory docs — missing Stock Ledger Entry
	for dt, amount_field in SLE_DOCTYPES:
		submitted = frappe.db.sql(
			f"""
			SELECT name, company, posting_date, owner, `{amount_field}` AS amount
			FROM `tab{dt}`
			WHERE docstatus = 1
			""",
			as_dict=True,
		)
		for d in submitted:
			has_sle = frappe.db.exists(
				"Stock Ledger Entry",
				{"voucher_type": dt, "voucher_no": d.name, "is_cancelled": 0},
			)
			if not has_sle:
				rows.append(
					{
						"doctype": dt,
						"voucher_no": d.name,
						"company": d.company,
						"posting_date": d.posting_date,
						"amount": flt(d.amount),
						"owner": d.owner,
						"missing": "Stock Ledger Entry",
						"category": "Inventory",
					}
				)

	return rows


@frappe.whitelist()
def summary():
	"""Return counts + totals by doctype & company."""
	rows = _collect_ghosts()
	by_doctype = {}
	by_company = {}
	for r in rows:
		by_doctype.setdefault(r["doctype"], {"count": 0, "amount": 0})
		by_doctype[r["doctype"]]["count"] += 1
		by_doctype[r["doctype"]]["amount"] += r["amount"] or 0

		by_company.setdefault(r["company"] or "(blank)", {"count": 0, "amount": 0})
		by_company[r["company"] or "(blank)"]["count"] += 1
		by_company[r["company"] or "(blank)"]["amount"] += r["amount"] or 0

	return {
		"total_count": len(rows),
		"total_amount": sum(r["amount"] or 0 for r in rows),
		"by_doctype": by_doctype,
		"by_company": by_company,
	}


@frappe.whitelist()
def export_excel():
	"""Download XLSX with 3 sheets: Summary, All Ghost Vouchers, By Company."""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from openpyxl.utils import get_column_letter

	if "System Manager" not in frappe.get_roles(frappe.session.user):
		frappe.throw("Only System Manager can export ghost vouchers.")

	rows = _collect_ghosts()
	rows.sort(key=lambda r: (r["company"] or "", r["doctype"], r["posting_date"] or ""))

	wb = Workbook()

	# ── Sheet 1: Summary ──
	ws = wb.active
	ws.title = "Summary"
	header_fill = PatternFill("solid", fgColor="1F4E78")
	header_font = Font(color="FFFFFF", bold=True)
	bold = Font(bold=True)

	ws["A1"] = "Ghost Voucher Report"
	ws["A1"].font = Font(bold=True, size=14)
	ws["A2"] = f"Generated: {frappe.utils.now_datetime().strftime('%Y-%m-%d %H:%M')}"
	ws["A3"] = f"Total ghost vouchers: {len(rows)}"
	ws["A3"].font = bold
	ws["A4"] = f"Total value: {sum(r['amount'] or 0 for r in rows):,.2f}"
	ws["A4"].font = bold

	ws["A6"] = "Breakdown by Document Type"
	ws["A6"].font = bold
	ws.append(["Doctype", "Category", "Count", "Total Amount"])
	for c in range(1, 5):
		cell = ws.cell(row=7, column=c)
		cell.fill = header_fill
		cell.font = header_font

	by_dt = {}
	for r in rows:
		key = (r["doctype"], r["category"])
		by_dt.setdefault(key, {"count": 0, "amount": 0})
		by_dt[key]["count"] += 1
		by_dt[key]["amount"] += r["amount"] or 0

	for (dt, cat), stats in sorted(by_dt.items(), key=lambda x: -x[1]["count"]):
		ws.append([dt, cat, stats["count"], round(stats["amount"], 2)])

	# Blank line, then by-company breakdown
	ws.append([])
	ws.append(["Breakdown by Company"])
	ws.cell(row=ws.max_row, column=1).font = bold
	ws.append(["Company", "Count", "Total Amount"])
	row_idx = ws.max_row
	for c in range(1, 4):
		cell = ws.cell(row=row_idx, column=c)
		cell.fill = header_fill
		cell.font = header_font

	by_co = {}
	for r in rows:
		co = r["company"] or "(blank)"
		by_co.setdefault(co, {"count": 0, "amount": 0})
		by_co[co]["count"] += 1
		by_co[co]["amount"] += r["amount"] or 0

	for co, stats in sorted(by_co.items(), key=lambda x: -x[1]["amount"]):
		ws.append([co, stats["count"], round(stats["amount"], 2)])

	for col in range(1, 5):
		ws.column_dimensions[get_column_letter(col)].width = 42

	# ── Sheet 2: All Ghost Vouchers ──
	ws2 = wb.create_sheet("All Ghost Vouchers")
	headers = [
		"Company",
		"Doctype",
		"Voucher No",
		"Posting Date",
		"Amount",
		"Category",
		"Missing Entry",
		"Owner",
	]
	ws2.append(headers)
	for c in range(1, len(headers) + 1):
		cell = ws2.cell(row=1, column=c)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center")

	for r in rows:
		ws2.append(
			[
				r["company"] or "",
				r["doctype"],
				r["voucher_no"],
				r["posting_date"].strftime("%Y-%m-%d") if r["posting_date"] else "",
				round(r["amount"] or 0, 2),
				r["category"],
				r["missing"],
				r["owner"] or "",
			]
		)

	widths = [42, 20, 28, 14, 16, 14, 20, 30]
	for i, w in enumerate(widths, start=1):
		ws2.column_dimensions[get_column_letter(i)].width = w
	ws2.freeze_panes = "A2"

	# ── Sheet 3: one sheet per company (top 6) ──
	companies = sorted(by_co.keys(), key=lambda c: -by_co[c]["amount"])
	for co in companies:
		safe = (co or "Unknown")[:25].replace("/", "-").replace("\\", "-").replace(":", "-")
		ws_co = wb.create_sheet(safe)
		ws_co.append(headers)
		for c in range(1, len(headers) + 1):
			cell = ws_co.cell(row=1, column=c)
			cell.fill = header_fill
			cell.font = header_font
		for r in rows:
			if (r["company"] or "(blank)") != co:
				continue
			ws_co.append(
				[
					r["company"] or "",
					r["doctype"],
					r["voucher_no"],
					r["posting_date"].strftime("%Y-%m-%d") if r["posting_date"] else "",
					round(r["amount"] or 0, 2),
					r["category"],
					r["missing"],
					r["owner"] or "",
				]
			)
		for i, w in enumerate(widths, start=1):
			ws_co.column_dimensions[get_column_letter(i)].width = w
		ws_co.freeze_panes = "A2"

	# ── Stream back as downloadable XLSX ──
	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	filename = f"Ghost_Vouchers_{frappe.utils.nowdate()}.xlsx"
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "binary"
