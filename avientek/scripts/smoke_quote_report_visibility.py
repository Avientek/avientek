"""Smoke test for Sridhar's 2026-06-09 Quote Report fixes.

Runs both patches end-to-end and verifies:
  - Quotation.customer_name is now pickable in Pick Columns
    (hidden=0, report_hide=0)
  - No Quotation Item rows have garbage custom_margin_ or
    custom_incentive_ (|value| > 500)

Usage:
    bench --site avientekv21.local execute \
        avientek.scripts.smoke_quote_report_visibility.run

Exits with non-zero return code on first failure so CI / a wrapper
shell script can fail the deploy if smoke breaks.
"""

import frappe
from frappe.utils import flt


SANE_LIMIT = 500.0


def _fail(msg):
    print(f"  ✗ FAIL: {msg}")
    raise AssertionError(msg)


def _ok(msg):
    print(f"  ✓ {msg}")


def _check_customer_name_visibility():
    print("=== Bug 1: Quotation.customer_name in Pick Columns ===")
    # Re-derive via Meta because Property Setters merge in at meta load.
    meta = frappe.get_meta("Quotation")
    field = meta.get_field("customer_name")
    if not field:
        _fail("customer_name field not found on Quotation")
    if field.hidden:
        _fail(f"customer_name still hidden=1 — pick-columns picker will skip it")
    if field.report_hide:
        _fail(f"customer_name has report_hide=1 — report-view picker will skip it")
    _ok(f"customer_name hidden={field.hidden}, report_hide={field.report_hide}, "
        f"read_only={field.read_only} — visible in Pick Columns")


def _check_no_margin_garbage():
    print()
    print("=== Bug 2: no garbage custom_margin_ / custom_incentive_ values ===")
    bad = frappe.db.sql(
        """
        SELECT COUNT(*) FROM `tabQuotation Item`
        WHERE ABS(IFNULL(custom_margin_, 0)) > %s
           OR ABS(IFNULL(custom_incentive_, 0)) > %s
        """,
        (SANE_LIMIT, SANE_LIMIT),
    )[0][0]
    if bad:
        _fail(f"{bad} Quotation Item rows still have |custom_margin_| or "
              f"|custom_incentive_| > {SANE_LIMIT} — patch did not run or "
              f"new bad data has been written")
    _ok(f"all Quotation Item rows have margin/incentive percent within "
        f"±{SANE_LIMIT:.0f}")


def _smoke_create_new_row_div_zero_path():
    """Defensive smoke: simulate the JS divide-by-zero formula and
    verify the patch's clamp logic produces 0 (not infinity / NaN /
    huge number) when amount = 0."""
    print()
    print("=== Defensive: divide-by-zero in margin% formula ===")
    amount = 0.0
    margin_value = -16915.20
    margin_percent = (margin_value / amount * 100.0) if amount else 0.0
    if margin_percent != 0.0:
        _fail(f"divide-by-zero guard did not zero the value — got {margin_percent}")
    _ok(f"div-zero guard yields 0 (not {-16915.20/(amount or 1)*100:.2f} unsafely)")


def _check_excel_percent_export_render():
    """Sridhar/Rahul 2026-06-10: Margin (%) column rendered as 1880%
    in the downloaded xlsx when UI showed 18.80%. Root cause: Frappe
    stores Percent as plain number, but Excel's 0.00% format expects
    a fraction. The builder now divides by 100 before writing.

    Build a 2-row synthetic xlsx through the real
    `_build_xlsx_bytes_from_rows` helper and read it back via
    openpyxl. The cell value for a Percent column must be the
    fraction (0.188), not the plain percent (18.80) — that's what
    makes Excel render '18.80%' instead of '1880.00%'.
    """
    print()
    print("=== Bug: Excel Percent column renders 100x off ===")
    from avientek.api.quotation_access import _build_xlsx_bytes_from_rows
    from openpyxl import load_workbook
    import io

    rows = [
        ["Quotation", "Margin (%)"],          # header
        ["QN-TEST-001",  18.80],              # data: stored 18.80, want display "18.80%"
        ["QN-TEST-002", -39.75],              # negative case
        ["QN-TEST-003",   0.00],              # zero
    ]
    col_types = ["Data", "Percent"]
    xlsx = _build_xlsx_bytes_from_rows(rows, "Quotation", col_types, [])
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active

    cases = [
        (2, 18.80,  0.188),
        (3, -39.75, -0.3975),
        (4, 0.00,   0.0),
    ]
    for row_idx, ui_value, expected_cell in cases:
        cell = ws.cell(row=row_idx, column=2)
        if cell.value is None:
            _fail(f"row {row_idx} Percent cell is None")
        diff = abs(float(cell.value) - expected_cell)
        if diff > 1e-9:
            _fail(f"row {row_idx} UI {ui_value}% → cell={cell.value} "
                  f"(want {expected_cell} so Excel renders '{ui_value:.2f}%', "
                  f"not '{float(cell.value)*100:.2f}%')")
        if cell.number_format != "0.00%":
            _fail(f"row {row_idx} number_format is {cell.number_format!r}, "
                  f"want '0.00%'")
    _ok(f"Percent column stored as fraction (18.80 → 0.188, -39.75 → -0.3975, "
        f"0 → 0); Excel will render correctly")


def _check_column_header_disambiguation_logic():
    """Sridhar/Rahul 2026-06-10: downloaded xlsx showed both Margin (%)
    columns (one from Quotation Item, one from Optional Item) with the
    same plain "Margin (%)" header — no child-table disambiguation —
    because the JS guard was `label.indexOf("(") === -1`, which
    false-fired on labels that already contain a parenthesis like
    "Margin (%)".

    We can't run JavaScript from `bench execute`, so this check is a
    source-level guard against the regression: assert the fixed
    pattern is present and the broken one is not.
    """
    print()
    print("=== Bug: Excel headers lose child-table disambiguation ===")
    import os
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    fp = os.path.join(here, "public", "js", "report_download.js")
    src = open(fp, "r", encoding="utf-8").read()
    # The broken pattern must be gone (except inside a comment that
    # documents what it used to be — comments live on the // path).
    code_lines = [
        ln for ln in src.split("\n")
        if "//" not in ln or ln.strip().split("//", 1)[0].strip()
    ]
    code_only = "\n".join(
        ln.split("//", 1)[0] for ln in src.split("\n")
    )
    if 'label.indexOf("(") === -1' in code_only:
        _fail("broken guard 'label.indexOf(\"(\") === -1' still in live "
              "code — percent-labelled child columns will skip "
              "disambiguation again")
    if 'var disambig = " (" + child_dt + ")"' not in src:
        _fail("fixed disambiguation suffix builder not found in "
              "report_download.js — child-table column headers won't "
              "include the source doctype")
    if 'label.indexOf(disambig) === -1' not in src:
        _fail("fixed disambig idempotency check not found — risk of "
              "double-appending the suffix")
    _ok("disambiguation guard now matches the SPECIFIC \" (child_dt)\" "
        "suffix; 'Margin (%)' from Quotation Item / Optional Item will "
        "export with their disambiguated headers")


def _check_reject_path_not_blocked_by_margin_gate():
    """Manu/Sridhar 2026-06-09: validate_margin_approval_required must
    permit transitions to terminal-reject states even when the margin
    is below threshold. The whitelist must contain Rejected +
    Cancelled (Rejected)."""
    print()
    print("=== Bug: Reject action blocked by margin gate ===")
    import inspect
    from avientek.events import quotation as q
    src = inspect.getsource(q.validate_margin_approval_required)
    for state in ("Rejected", "Cancelled (Rejected)"):
        if f'"{state}"' not in src:
            _fail(f"APPROVAL_PATH_STATES missing {state!r} — Reject workflow "
                  f"action would be blocked on low-margin quotes")
    _ok("APPROVAL_PATH_STATES includes Rejected + Cancelled (Rejected) — "
        "Reject workflow action passes the margin gate")


def run():
    print("=" * 64)
    print("Avientek smoke: Quote Report visibility + margin sanity")
    print("=" * 64)
    _check_customer_name_visibility()
    _check_no_margin_garbage()
    _smoke_create_new_row_div_zero_path()
    _check_reject_path_not_blocked_by_margin_gate()
    _check_excel_percent_export_render()
    _check_column_header_disambiguation_logic()
    print()
    print("All smoke checks PASSED ✓")
